from __future__ import annotations

import os
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from config import Config


try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


@dataclass
class CaseInfo:
    index: int
    excel_row_number: int
    raw_row: Dict[str, Any]
    settings: Dict[str, Any]
    mesh_file: Path
    case_basename: str
    journal_file: Path
    transcript_file: Path
    output_case_data: Path
    monitor_file: Path
    console_log_file: Path


class Solver:
    """同一 journal 中：读网格 -> 稳态 -> 切换非稳态 -> 非稳态"""

    def __init__(self, config: Config):
        self.config = config

    def load_cases_from_excel(self) -> List[CaseInfo]:
        excel_path = self.config.solve_table_path
        if not excel_path.exists():
            raise FileNotFoundError(f"求解参数表不存在: {excel_path}")

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[self.config.solve_sheet_name] if self.config.solve_sheet_name else wb[wb.sheetnames[0]]

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not header:
            raise ValueError("solve_parameters.xlsx 第一行表头为空。")

        rows: List[CaseInfo] = []
        case_counter = 0

        for excel_row_number, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == "" for v in row_cells):
                continue

            case_counter += 1

            raw_row = {
                str(header[i]).strip(): row_cells[i]
                for i in range(len(header))
                if header[i] is not None
            }

            settings = dict(self.config.default_case_settings)
            for key, value in raw_row.items():
                if key in self.config.recognized_columns and value is not None and str(value).strip() != "":
                    settings[key] = value

            # 兼容旧列名：number_of_iterations -> steady_iterations
            if (
                ("steady_iterations" not in raw_row or raw_row.get("steady_iterations") in (None, ""))
                and ("number_of_iterations" in raw_row and raw_row.get("number_of_iterations") not in (None, ""))
            ):
                settings["steady_iterations"] = raw_row["number_of_iterations"]

            mesh_value = raw_row.get("mesh_file", raw_row.get("mesh"))
            mesh_file = self.config.resolve_mesh_file(mesh_value)

            case_basename = self.config.get_case_basename(case_counter, raw_row)

            journal_file = self.config.journal_dir / f"solve{case_counter}.jou"
            transcript_file = self.config.transcript_dir / f"solve{case_counter}.trn"
            output_case_data = self.config.solve_dir / f"solve{case_counter}.cas.h5"
            monitor_file = self.config.result_dir / f"force_moment_history_solve{case_counter}.out"
            console_log_file = self.config.result_dir / f"solve{case_counter}.console.log"

            rows.append(
                CaseInfo(
                    index=case_counter,
                    excel_row_number=excel_row_number,
                    raw_row=raw_row,
                    settings=settings,
                    mesh_file=mesh_file,
                    case_basename=case_basename,
                    journal_file=journal_file,
                    transcript_file=transcript_file,
                    output_case_data=output_case_data,
                    monitor_file=monitor_file,
                    console_log_file=console_log_file,
                )
            )

        if not rows:
            raise ValueError("solve_parameters.xlsx 中没有可用的求解工况。")

        return rows

    def get_case_count(self) -> int:
        return len(self.load_cases_from_excel())

    def get_case_by_index(self, case_index: int) -> CaseInfo:
        cases = self.load_cases_from_excel()
        if case_index < 1 or case_index > len(cases):
            raise IndexError(f"工况编号越界: {case_index}，当前共有 {len(cases)} 个有效工况。")
        return cases[case_index - 1]

    def _to_float(self, value: Any, name: str) -> float:
        try:
            return float(value)
        except Exception as exc:
            raise ValueError(f"参数 {name} 不能转换为浮点数: {value}") from exc

    def _to_int(self, value: Any, name: str) -> int:
        try:
            return int(float(value))
        except Exception as exc:
            raise ValueError(f"参数 {name} 不能转换为整数: {value}") from exc

    def _to_bool(self, value: Any, default: bool = True) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "on"}:
            return True
        if text in {"0", "false", "no", "n", "off"}:
            return False
        return bool(value)

    def build_journal_text(self, case: CaseInfo) -> str:
        s = case.settings

        model = str(s.get("model", "transition-sst")).strip().lower()
        if model != "transition-sst":
            raise ValueError(f"暂不支持的 model: {model}")

        inlet_name = str(s.get("inlet_name", self.config.inlet_name)).strip()
        inlet_velocity = self._to_float(s.get("inlet_velocity", 1.0), "inlet_velocity")
        viscosity = self._to_float(s.get("viscosity", 1.7894e-5), "viscosity")
        pv_coupling = self._to_int(s.get("pv_coupling", 24), "pv_coupling")

        steady_iterations = self._to_int(s.get("steady_iterations", 150), "steady_iterations")
        dt = self._to_float(s.get("time_step", 0.05), "time_step")
        n_steps = self._to_int(s.get("number_of_time_steps", 600), "number_of_time_steps")
        max_iter_per_step = self._to_int(s.get("max_iter_per_time_step", 10), "max_iter_per_time_step")

        intensity = self._to_float(
            s.get("inlet_turbulence_intensity", self.config.inlet_turbulence_intensity),
            "inlet_turbulence_intensity",
        )
        vis_ratio = self._to_float(
            s.get("inlet_turbulence_viscosity_ratio", self.config.inlet_turbulence_viscosity_ratio),
            "inlet_turbulence_viscosity_ratio",
        )

        write_inst = self.config.bool_to_yes_no(
            self._to_bool(s.get("write_instantaneous_values", True), True)
        )
        output_case_data = self._to_bool(s.get("output_case_data", True), True)

        zone_block = " ".join(self.config.wall_zones)
        mom_center = self.config.moment_center

        lines: List[str] = []
        lines.append("; =========================")
        lines.append(f"; Fluent continuous steady -> unsteady journal: solve{case.index}")
        lines.append(f"; excel_row = {case.excel_row_number}, case_name = {case.case_basename}")
        lines.append("; auto generated by solver.py")
        lines.append("; =========================")
        lines.append("")

        lines.append(f'/file/start-transcript "{case.transcript_file.as_posix()}"')
        lines.append(f'/file/read-case "{case.mesh_file.as_posix()}"')
        lines.append("")

        lines.append("; 1) 模型：先稳态设置")
        lines.append("/define/models/viscous/transition-sst yes")
        lines.append("")

        lines.append("; 2) 材料")
        lines.append("/define/materials/change-create air air")
        lines.extend([
            "no",
            "no",
            "no",
            "yes",
            "constant",
            f"{viscosity:.12g}",
            "no",
            "no",
            "no",
            "",
        ])

        lines.append("; 3) 边界条件：入口")
        lines.append(f"/define/boundary-conditions/velocity-inlet {inlet_name}")
        lines.extend([
            "no",
            "no",
            "yes",
            "yes",
            "no",
            f"{inlet_velocity:.12g}",
            "no",
            "0",
            "no",
            "no",
            "yes",
            "no",
            "1",
            f"{intensity:.12g}",
            f"{vis_ratio:.12g}",
            "",
        ])

        lines.append("; 4) 求解方法")
        lines.append(f"/solve/set/p-v-coupling {pv_coupling}")
        lines.append("")

        lines.append("; 5) 报告定义：力")
        for name, vec in (("fx", (1, 0, 0)), ("fy", (0, 1, 0)), ("fz", (0, 0, 1))):
            lines.extend([
                "/solve/report-definitions/add",
                name,
                "force",
                "force-vector",
                str(vec[0]),
                str(vec[1]),
                str(vec[2]),
                "thread-names",
                zone_block,
                "",
                "per-zone",
                "no",
                "q",
                "",
            ])

        lines.append("; 6) 报告定义：力矩")
        for name, axis in (("momx", (1, 0, 0)), ("momy", (0, 1, 0)), ("momz", (0, 0, 1))):
            lines.extend([
                "/solve/report-definitions/add",
                name,
                "moment",
                "scaled",
                "no",
                "mom-center",
                f"{mom_center[0]:.12g}",
                f"{mom_center[1]:.12g}",
                f"{mom_center[2]:.12g}",
                "mom-axis",
                str(axis[0]),
                str(axis[1]),
                str(axis[2]),
                "thread-names",
                zone_block,
                "",
                "per-zone",
                "no",
                "q",
                "",
            ])

        lines.append("; 7) 初始化")
        init_method = str(s.get("initialize_method", "hyb-initialization")).strip()
        if init_method != "hyb-initialization":
            raise ValueError(f"暂不支持的 initialize_method: {init_method}")
        lines.append("/solve/initialize/hyb-initialization")
        lines.append("")

        lines.append("; 8) 新建监控输出文件")
        lines.extend([
            "/solve/report-files/add/force-moment-file",
            "file-name",
            f'"{case.monitor_file.as_posix()}"',
            "report-defs",
            "fx",
            "fy",
            "fz",
            "momx",
            "momy",
            "momz",
            "",
            "print?",
            "yes",
            "plot?",
            "no",
            "write?",
            "yes",
            "append?",
            "no",
            "write-case?",
            write_inst,
            "q",
            "",
        ])

        lines.append("; 9) 先做稳态")
        lines.append(f"/solve/iterate {steady_iterations}")
        lines.append("")

        lines.append("; 10) 切换到非稳态")
        lines.append("/define/models/unsteady-1st-order yes")
        lines.extend([
            f"/solve/set/transient-controls/time-step-size {dt:.12g}",
            "/solve/set/transient-controls/fixed-user-specified yes",
            f"/solve/set/transient-controls/number-of-time-steps {n_steps}",
            f"/solve/set/transient-controls/max-iterations-per-time-step {max_iter_per_step}",
            "/solve/set/transient-controls/extrapolate-vars? yes",
            "",
        ])

        lines.append("; 11) 直接继续非稳态")
        lines.append(f"/solve/dual-time-iterate {n_steps} {max_iter_per_step}")
        lines.append("")

        if output_case_data:
            lines.append("; 12) 保存最终 case/data")
            lines.append(f'/file/write-case-data "{case.output_case_data.as_posix()}"')
            lines.append("")

        lines.append("; 13) 退出")
        lines.extend([
            "/file/stop-transcript",
            "exit",
            "yes",
            "",
        ])

        return "\n".join(lines)

    def write_journal(self, case: CaseInfo) -> None:
        text = self.build_journal_text(case)
        case.journal_file.parent.mkdir(parents=True, exist_ok=True)
        case.journal_file.write_text(text, encoding="utf-8", newline="\n")
        print(f"已生成 journal: {case.journal_file}")

    def _resolve_processor_count(self, case: CaseInfo) -> int | None:
        s = case.settings

        value = s.get("processor_count")
        if value is not None and str(value).strip() != "":
            return self._to_int(value, "processor_count")

        if self.config.processor_count is not None:
            return self._to_int(self.config.processor_count, "processor_count")

        slurm_ntasks = os.environ.get("SLURM_NTASKS")
        if slurm_ntasks:
            return self._to_int(slurm_ntasks, "SLURM_NTASKS")

        return None

    def build_fluent_command(self, case: CaseInfo) -> List[str]:
        s = case.settings
        processor_count = self._resolve_processor_count(case)

        dimension = str(s.get("dimension", self.config.dimension)).strip().lower()
        precision = str(s.get("precision", self.config.precision)).strip().lower()

        if dimension == "3d" and precision == "dp":
            mode = "3ddp"
        elif dimension == "3d" and precision in {"sp", "single"}:
            mode = "3d"
        elif dimension == "2d" and precision == "dp":
            mode = "2ddp"
        elif dimension == "2d" and precision in {"sp", "single"}:
            mode = "2d"
        else:
            raise ValueError(f"不支持的 Fluent 维度/精度组合: dimension={dimension}, precision={precision}")

        cmd = [self.config.fluent_path, mode]

        if os.environ.get("SLURM_JOB_ID"):
            cmd.append("-slurm")

        if processor_count is not None:
            cmd.append(f"-t{processor_count}")

        if self.config.headless:
            cmd.append("-g")

        cmd.extend(["-i", str(case.journal_file)])
        return cmd

    def run_case(self, case: CaseInfo, dry_run: bool = False) -> int | None:
        self.write_journal(case)
        cmd = self.build_fluent_command(case)

        print("=" * 72)
        print(f"开始工况 {case.index}: {case.case_basename}")
        print(f"Excel 行号     : {case.excel_row_number}")
        print(f"网格文件       : {case.mesh_file}")
        print(f"Journal        : {case.journal_file}")
        print(f"Transcript     : {case.transcript_file}")
        print(f"Monitor        : {case.monitor_file}")
        print(f"Case/Data输出  : {case.output_case_data}")
        print(f"Console日志    : {case.console_log_file}")
        print(f"并行核数       : {processor_count if (processor_count := self._resolve_processor_count(case)) is not None else '未显式指定'}")
        print("Fluent 命令:")
        print(" ".join(cmd))

        if dry_run:
            print("dry_run=True，本工况只生成 journal，不实际启动 Fluent。")
            return None

        with open(case.console_log_file, "w", encoding="utf-8", buffering=1) as log:
            process = subprocess.Popen(
                cmd,
                cwd=self.config.base_dir,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                universal_newlines=True,
            )

            assert process.stdout is not None
            for line in process.stdout:
                print(line, end="")
                log.write(line)

            process.wait()

        print(f"\n返回码: {process.returncode}")
        print(f"控制台日志已保存: {case.console_log_file}")
        print(f"transcript 路径: {case.transcript_file}")

        if process.returncode != 0:
            raise RuntimeError(
                f"Fluent 运行失败: solve{case.index}\n"
                f"返回码: {process.returncode}\n"
                f"请检查: {case.console_log_file}\n"
                f"以及 transcript: {case.transcript_file}"
            )

        return process.returncode

    def run_all(self, dry_run: bool = False) -> List[int | None]:
        cases = self.load_cases_from_excel()
        print(f"读取到 {len(cases)} 个求解工况。")
        results: List[int | None] = []
        for case in cases:
            results.append(self.run_case(case, dry_run=dry_run))
        return results

    def run_one_by_index(self, case_index: int, dry_run: bool = False) -> int | None:
        case = self.get_case_by_index(case_index)
        return self.run_case(case, dry_run=dry_run)
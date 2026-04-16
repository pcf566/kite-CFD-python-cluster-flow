from __future__ import annotations

import os
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from config import Config


try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


@dataclass
class UnsteadyCaseInfo:
    index: int
    excel_row_number: int
    raw_row: Dict[str, Any]
    settings: Dict[str, Any]
    steady_case_index: int
    steady_case_file: Path
    steady_data_file: Path
    journal_file: Path
    transcript_file: Path
    output_case_data_file: Path
    monitor_file: Path
    console_log_file: Path
    report_def_names: List[str] = field(default_factory=list)


class UnsteadySolver:
    """
    从稳态结果继续做非稳态：
    1) 读取 solve/solveN.cas.h5 + solve/solveN.dat.h5
    2) 切到非稳态
    3) 设置时间步参数
    4) 删除旧的 report-file / report-plot
    5) 根据现有 case 中的 report definition 名称，重建新的 monitor
    6) 进行 dual-time-iterate
    """

    def __init__(
        self,
        config: Config,
        excel_name: str = "unsteady.xlsx",
        sheet_name: str | None = None,
    ):
        self.config = config
        self.excel_name = excel_name
        self.sheet_name = sheet_name

    @property
    def excel_path(self) -> Path:
        candidates = [
            self.config.base_dir / self.excel_name,
            self.config.solve_dir / self.excel_name,
        ]
        for p in candidates:
            if p.exists():
                return p.resolve()
        return candidates[0].resolve()

    def _to_float(self, value: Any, name: str) -> float:
        try:
            return float(value)
        except Exception as exc:
            raise ValueError(f"参数 {name} 不能转换为浮点数: {value}") from exc

    def _to_int(self, value: Any, name: str) -> int:
        try:
            return int(float(value))
        except Exception as exc:
            raise ValueError(f"参数 {name} 不能转换为整数: {value}") from exc

    def _to_bool(self, value: Any, default: bool = True) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "on"}:
            return True
        if text in {"0", "false", "no", "n", "off"}:
            return False
        return bool(value)

    def load_cases_from_excel(self) -> List[UnsteadyCaseInfo]:
        excel_path = self.excel_path
        if not excel_path.exists():
            raise FileNotFoundError(
                f"非稳态参数表不存在: {excel_path}\n"
                f"请把 {self.excel_name} 放到项目根目录或 solve 目录中。"
            )

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb[wb.sheetnames[0]]

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not header:
            raise ValueError(f"{excel_path.name} 第一行表头为空。")

        required_columns = {
            "case",
        }
        header_set = {str(h).strip() for h in header if h is not None}
        missing = required_columns - header_set
        if missing:
            raise ValueError(
                f"{excel_path.name} 缺少必要列: {sorted(missing)}。\n"
                "至少应包含: case"
            )

        rows: List[UnsteadyCaseInfo] = []
        case_counter = 0

        for excel_row_number, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == "" for v in row_cells):
                continue

            case_counter += 1

            raw_row = {
                str(header[i]).strip(): row_cells[i]
                for i in range(len(header))
                if header[i] is not None
            }

            steady_case_index = self._to_int(raw_row.get("case"), "case")

            steady_solve_dir = raw_row.get("steady_solve_dir")
            if steady_solve_dir is not None and str(steady_solve_dir).strip() != "":
                steady_solve_dir = Path(str(steady_solve_dir).strip()).expanduser().resolve()
            else:
                steady_solve_dir = self.config.solve_dir

            steady_case_file = steady_solve_dir / f"solve{steady_case_index}.cas.h5"
            steady_data_file = steady_solve_dir / f"solve{steady_case_index}.dat.h5"

            if not steady_case_file.exists():
                raise FileNotFoundError(f"稳态 case 文件不存在: {steady_case_file}")
            if not steady_data_file.exists():
                raise FileNotFoundError(f"稳态 data 文件不存在: {steady_data_file}")

            settings = dict(self.config.default_case_settings)
            settings["time_type"] = "unsteady-1st-order"

            recognized_unsteady_columns = {
                "case",
                "case_name",
                "model",
                "pv_coupling",
                "time_step",
                "number_of_time_steps",
                "max_iter_per_time_step",
                "processor_count",
                "precision",
                "dimension",
                "write_instantaneous_values",
                "output_case_data",
                "output_basename",
                "append_monitor",
            }
            for key, value in raw_row.items():
                if key in recognized_unsteady_columns and value is not None and str(value).strip() != "":
                    settings[key] = value

            output_basename = raw_row.get("output_basename")
            if output_basename is None or str(output_basename).strip() == "":
                output_basename = f"solve{steady_case_index}_unsteady"

            output_basename = str(output_basename).strip()
            journal_file = self.config.journal_dir / f"{output_basename}.jou"
            transcript_file = self.config.transcript_dir / f"{output_basename}.trn"
            output_case_data_file = self.config.solve_dir / f"{output_basename}.cas.h5"
            monitor_file = self.config.result_dir / f"force_moment_history_{output_basename}.out"
            console_log_file = self.config.result_dir / f"{output_basename}.console.log"

            rows.append(
                UnsteadyCaseInfo(
                    index=case_counter,
                    excel_row_number=excel_row_number,
                    raw_row=raw_row,
                    settings=settings,
                    steady_case_index=steady_case_index,
                    steady_case_file=steady_case_file,
                    steady_data_file=steady_data_file,
                    journal_file=journal_file,
                    transcript_file=transcript_file,
                    output_case_data_file=output_case_data_file,
                    monitor_file=monitor_file,
                    console_log_file=console_log_file,
                )
            )

        if not rows:
            raise ValueError(f"{excel_path.name} 中没有可用的非稳态工况。")
        return rows

    def get_case_count(self) -> int:
        return len(self.load_cases_from_excel())

    def get_case_by_index(self, case_index: int) -> UnsteadyCaseInfo:
        cases = self.load_cases_from_excel()
        if case_index < 1 or case_index > len(cases):
            raise IndexError(f"工况编号越界: {case_index}，当前共有 {len(cases)} 个有效工况。")
        return cases[case_index - 1]

    def _steady_monitor_candidates(self, case: UnsteadyCaseInfo) -> List[Path]:
        idx = case.steady_case_index
        return [
            self.config.result_dir / f"force_moment_history{idx}.out",
            self.config.result_dir / f"force_moment_history_solve{idx}.out",
            self.config.result_dir / f"solve{idx}.out",
        ]

    def _extract_header_report_defs(self, text: str) -> List[str]:
        m = re.search(r'\(([^\n]*"[^"]+"[^\n]*)\)', text)
        if not m:
            return []
        quoted = re.findall(r'"([^"]+)"', m.group(0))
        defs: List[str] = []
        for item in quoted:
            name = item.strip()
            if name in {"Time Step", "flow-time"}:
                continue
            defs.append(name)
        return defs

    def _detect_from_existing_monitor(self, case: UnsteadyCaseInfo) -> List[str]:
        for p in self._steady_monitor_candidates(case):
            if not p.exists():
                continue
            try:
                text = p.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            defs = self._extract_header_report_defs(text)
            if defs:
                return defs
        return []

    def _build_probe_command(self, journal_path: Path, processor_count: int) -> List[str]:
        dimension = str(self.config.dimension).strip().lower()
        precision = str(self.config.precision).strip().lower()
        if dimension == "3d" and precision == "dp":
            mode = "3ddp"
        elif dimension == "3d" and precision in {"sp", "single"}:
            mode = "3d"
        elif dimension == "2d" and precision == "dp":
            mode = "2ddp"
        elif dimension == "2d" and precision in {"sp", "single"}:
            mode = "2d"
        else:
            raise ValueError(f"不支持的 Fluent 维度/精度组合: dimension={dimension}, precision={precision}")

        cmd = [self.config.fluent_path, mode]
        if os.environ.get("SLURM_JOB_ID"):
            cmd.append("-slurm")
        cmd.append(f"-t{processor_count}")
        if self.config.headless:
            cmd.append("-g")
        cmd.extend(["-i", str(journal_path)])
        return cmd

    def _detect_by_probe(self, case: UnsteadyCaseInfo) -> List[str]:
        processor_count = 1
        probe_defs_file = self.config.result_dir / f"probe_report_defs_solve{case.steady_case_index}.txt"
        probe_log_file = self.config.result_dir / f"probe_report_defs_solve{case.steady_case_index}.console.log"
        self.config.result_dir.mkdir(parents=True, exist_ok=True)

        probe_lines = [
            f'/file/read-case "{case.steady_case_file.as_posix()}"',
            f'/file/read-data "{case.steady_data_file.as_posix()}"',
            '/solve/report-definitions/list',
            'exit',
            'yes',
            '',
        ]

        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            suffix=".jou",
            prefix=f"probe_report_defs_solve{case.steady_case_index}_",
            dir=str(self.config.journal_dir),
            delete=False,
            newline="\n",
        ) as tmp:
            tmp.write("\n".join(probe_lines))
            probe_journal = Path(tmp.name)

        cmd = self._build_probe_command(probe_journal, processor_count=processor_count)
        completed = subprocess.run(
            cmd,
            cwd=self.config.base_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
        probe_log_file.write_text(completed.stdout, encoding="utf-8", newline="\n")

        names = sorted(set(re.findall(r'\b(?:fx|fy|fz|mx|my|mz|momx|momy|momz)\b', completed.stdout)))
        if names:
            probe_defs_file.write_text("\n".join(names) + "\n", encoding="utf-8", newline="\n")

        try:
            probe_journal.unlink(missing_ok=True)
        except Exception:
            pass

        return names

    def detect_report_definitions(self, case: UnsteadyCaseInfo, dry_run: bool = False) -> List[str]:
        if case.report_def_names:
            return case.report_def_names

        defs = self._detect_from_existing_monitor(case)
        if defs:
            case.report_def_names = defs
            return defs

        if not dry_run:
            defs = self._detect_by_probe(case)
            if defs:
                case.report_def_names = defs
                return defs

        # 保底：至少保留力，力矩按最常见的 mx/my/mz 估计
        fallback = ["fx", "fy", "fz",  "momx", "momy", "momz"]
        case.report_def_names = fallback
        return fallback

    def _normalize_report_defs(self, defs: List[str]) -> List[str]:
        preferred_order = ["fx", "fy", "fz", "mx", "my", "mz", "momx", "momy", "momz"]
        present = set(defs)
        ordered = [name for name in preferred_order if name in present]
        # 只允许一种力矩命名风格，优先沿用检测到的风格
        if any(name in present for name in ["momx", "momy", "momz"]):
            ordered = [name for name in ordered if name not in {"mx", "my", "mz"}]
        elif any(name in present for name in ["mx", "my", "mz"]):
            ordered = [name for name in ordered if name not in {"momx", "momy", "momz"}]
        if not any(name in ordered for name in ["fx", "fy", "fz"]):
            ordered = ["fx", "fy", "fz"] + ordered
        return ordered

    def build_journal_text(self, case: UnsteadyCaseInfo) -> str:
        s = case.settings

        model = str(s.get("model", "transition-sst")).strip().lower()
        if model != "transition-sst":
            raise ValueError(f"暂不支持的 model: {model}")

        pv_coupling = self._to_int(s.get("pv_coupling", 24), "pv_coupling")
        dt = self._to_float(s.get("time_step"), "time_step")
        n_steps = self._to_int(s.get("number_of_time_steps"), "number_of_time_steps")
        max_iter_per_step = self._to_int(s.get("max_iter_per_time_step"), "max_iter_per_time_step")
        output_case_data = self._to_bool(s.get("output_case_data"), True)

        report_defs = self._normalize_report_defs(case.report_def_names)

        lines: List[str] = []
        lines.append("; =========================")
        lines.append(f"; Fluent 2022 R1 unsteady continuation: case {case.index}")
        lines.append(f"; excel_row = {case.excel_row_number}, steady_case = solve{case.steady_case_index}")
        lines.append("; auto generated by unsteady_solver.py")
        lines.append("; =========================")
        lines.append("")
        lines.append(f'/file/start-transcript "{case.transcript_file.as_posix()}"')
        lines.append(f'/file/read-case "{case.steady_case_file.as_posix()}"')
        lines.append(f'/file/read-data "{case.steady_data_file.as_posix()}"')
        lines.append("")

        lines.append("; 1) 从稳态切换为非稳态")
        lines.append("/define/models/unsteady-1st-order yes")
        lines.append("/define/models/viscous/transition-sst yes")
        lines.append(f"/solve/set/p-v-coupling {pv_coupling}")
        lines.append("")

        lines.append("; 2) 瞬态控制")
        lines.extend([
            f"/solve/set/transient-controls/time-step-size {dt:.12g}",
            "/solve/set/transient-controls/fixed-user-specified yes",
            f"/solve/set/transient-controls/number-of-time-steps {n_steps}",
            f"/solve/set/transient-controls/max-iterations-per-time-step {max_iter_per_step}",
            "/solve/set/transient-controls/extrapolate-vars? yes",
            "",
        ])

        lines.append("; 3) 清空旧 report-file / report-plot，保留已有 report-definition")
        lines.extend([
            "/solve/report-files/delete-all",
            "yes",
            "/solve/report-plots/delete-all",
            "yes",
            "",
        ])

        lines.append("; 4) 新建非稳态 monitor 文件")
        lines.extend([
            "/solve/report-files/add",
            "force-moment-file",
            "report-file",
            "file-name",
            f'"{case.monitor_file.as_posix()}"',
            "report-defs",
            *report_defs,
            "",
            "print?",
            "yes",
            "q",
            "",
        ])

        lines.append("; 5) 正式计算")
        lines.append(f"/solve/dual-time-iterate {n_steps} {max_iter_per_step}")
        lines.append("")

        if output_case_data:
            lines.append("; 6) 保存新的 case/data")
            lines.append(f'/file/write-case-data "{case.output_case_data_file.as_posix()}"')
            lines.append("")

        lines.append("; 7) 退出")
        lines.extend([
            "/file/stop-transcript",
            "exit",
            "yes",
            "",
        ])

        return "\n".join(lines)

    def write_journal(self, case: UnsteadyCaseInfo) -> None:
        text = self.build_journal_text(case)
        case.journal_file.parent.mkdir(parents=True, exist_ok=True)
        case.journal_file.write_text(text, encoding="utf-8", newline="\n")
        print(f"已生成 journal: {case.journal_file}")

    def build_fluent_command(self, case: UnsteadyCaseInfo) -> List[str]:
        s = case.settings
        processor_count = self._to_int(
            s.get("processor_count", self.config.processor_count),
            "processor_count",
        )

        dimension = str(s.get("dimension", self.config.dimension)).strip().lower()
        precision = str(s.get("precision", self.config.precision)).strip().lower()

        if dimension == "3d" and precision == "dp":
            mode = "3ddp"
        elif dimension == "3d" and precision in {"sp", "single"}:
            mode = "3d"
        elif dimension == "2d" and precision == "dp":
            mode = "2ddp"
        elif dimension == "2d" and precision in {"sp", "single"}:
            mode = "2d"
        else:
            raise ValueError(f"不支持的 Fluent 维度/精度组合: dimension={dimension}, precision={precision}")

        cmd = [self.config.fluent_path, mode]

        if os.environ.get("SLURM_JOB_ID"):
            cmd.append("-slurm")

        cmd.append(f"-t{processor_count}")

        if self.config.headless:
            cmd.append("-g")

        cmd.extend(["-i", str(case.journal_file)])
        return cmd

    def run_case(self, case: UnsteadyCaseInfo, dry_run: bool = False) -> int | None:
        detected_defs = self.detect_report_definitions(case, dry_run=dry_run)
        print(f"检测到 report definitions: {detected_defs}")

        self.write_journal(case)
        cmd = self.build_fluent_command(case)

        print("=" * 72)
        print(f"开始非稳态工况 {case.index}")
        print(f"Excel 行号      : {case.excel_row_number}")
        print(f"稳态来源 case   : {case.steady_case_file}")
        print(f"稳态来源 data   : {case.steady_data_file}")
        print(f"ReportDefs      : {case.report_def_names}")
        print(f"Journal         : {case.journal_file}")
        print(f"Transcript      : {case.transcript_file}")
        print(f"Monitor         : {case.monitor_file}")
        print(f"Case/Data输出   : {case.output_case_data_file}")
        print(f"Console日志     : {case.console_log_file}")
        print("Fluent 命令:")
        print(" ".join(cmd))

        if dry_run:
            print("dry_run=True，本工况只生成 journal，不实际启动 Fluent。")
            return None

        with open(case.console_log_file, "w", encoding="utf-8", buffering=1) as log:
            process = subprocess.Popen(
                cmd,
                cwd=self.config.base_dir,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                universal_newlines=True,
            )

            assert process.stdout is not None
            for line in process.stdout:
                print(line, end="")
                log.write(line)

            process.wait()

        print(f"\n返回码: {process.returncode}")
        print(f"控制台日志已保存: {case.console_log_file}")
        print(f"transcript 路径: {case.transcript_file}")

        if process.returncode != 0:
            raise RuntimeError(
                f"Fluent 运行失败: 非稳态工况 {case.index}\n"
                f"返回码: {process.returncode}\n"
                f"请检查: {case.console_log_file}\n"
                f"以及 transcript: {case.transcript_file}"
            )

        return process.returncode

    def run_all(self, dry_run: bool = False) -> List[int | None]:
        cases = self.load_cases_from_excel()
        print(f"读取到 {len(cases)} 个非稳态工况。")
        results: List[int | None] = []
        for case in cases:
            results.append(self.run_case(case, dry_run=dry_run))
        return results

    def run_one_by_index(self, case_index: int, dry_run: bool = False) -> int | None:
        case = self.get_case_by_index(case_index)
        return self.run_case(case, dry_run=dry_run)
